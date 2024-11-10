import openpyxl
import pandas as pd
import keyboard
from datetime import datetime

while True:
    # Create an Excel file if it doesn’t exist
    file_name = "sales_data.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales Data"
        # Adding headers
        sheet.append(["Date", "Product", "Quantity", "Price", "Total Sales"])
        workbook.save(file_name)

    # Function to add data
    def add_sales_data(date, product, quantity, price):
        total_sales = quantity * price
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["Sales Data"]
        sheet.append([date, product, quantity, price, total_sales])
        workbook.save(file_name)
        print("Data added successfully!")

    # Collect input data
    try:
        today_date = datetime.now().strftime("%Y-%m-%d")
        product_name = input("Enter product name: ")
        quantity_sold = int(input("Enter quantity sold: "))
        price_per_unit = float(input("Enter price per unit: "))
        try:
            add_sales_data(today_date, product_name, quantity_sold, price_per_unit)  #addds data to excel
        except:
            print(f'Close the file {file_name} and try again')
    except:
        print('Invalid Entry')

    print("Press any key to continue or press '0' to exit.")
    event = keyboard.read_event(suppress=True)  # Capture key press and suppress echo
    key = event.name  # Get the name of the key pressed
    if key == '0':  # Exit if '0' is pressed
        print("Exiting...")
        break
